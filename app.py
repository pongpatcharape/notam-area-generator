import io
import os
import re
import zipfile
import folium
import geopandas as gpd
import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from pyproj import Geod
import simplekml
import streamlit as st
from streamlit_folium import st_folium

# ⚙️ 1. SET UP PAGE & FIT-TO-SCREEN CSS
st.set_page_config(
    page_title="NOTAM AREA GENERATOR",
    page_icon="✈️",
    layout="wide",
    initial_sidebar_state="expanded",
)

st.markdown(
    """
    <style>
    /* 1. ล็อค Scrollbar และปรับ Padding หน้าจอ */
    html, body, [data-testid="stAppViewContainer"] {
        overflow: hidden !important;
    }
    #MainMenu {visibility: hidden;}
    footer {visibility: hidden;}
    header {visibility: hidden;}
    
    .main .block-container {
        padding-top: 0.5rem !important;
        padding-bottom: 0rem !important;
        padding-left: 1rem !important;
        padding-right: 1rem !important;
        max-width: 100% !important;
    }

    /* 2. Header Layout */
    .header-title {
        font-size: 1.4rem !important;
        font-weight: 800;
        margin: 0 !important;
        padding: 0 !important;
        color: #f8fafc;
        line-height: 1.2;
    }
    .header-desc {
        font-size: 0.8rem !important;
        color: #94a3b8;
        margin-top: 2px !important;
        margin-bottom: 0 !important;
    }

    /* 3. Metric Cards 2x2 ด้านขวา (แบบย่อสเปคกะทัดรัด) */
    .metric-grid {
        display: grid;
        grid-template-columns: 1fr 1fr;
        gap: 4px;
    }
    .metric-card-mini {
        background-color: #1e293b;
        border: 1px solid #334155;
        border-radius: 6px;
        padding: 4px 8px;
        text-align: center;
    }
    .metric-card-mini h5 {
        margin: 0;
        color: #94a3b8;
        font-size: 0.65rem;
        font-weight: 600;
        text-transform: uppercase;
        letter-spacing: 0.3px;
    }
    .metric-card-mini p {
        margin: 1px 0 0 0;
        color: #f8fafc;
        font-size: 0.8rem;
        font-weight: 700;
        white-space: nowrap;
        overflow: hidden;
        text-overflow: ellipsis;
    }

    /* 4. Streamlit Tabs & Buttons Compact */
    .stTabs [data-baseweb="tab-list"] {
        gap: 8px;
    }
    .stTabs [data-baseweb="tab"] {
        height: 32px;
        padding-top: 0px;
        padding-bottom: 0px;
        font-size: 0.85rem;
    }
    
    .stButton > button {
        width: 100%;
        background-color: #0d6efd;
        color: white;
        font-weight: 600;
        border-radius: 6px;
        padding: 0.4rem 0.8rem;
        border: none;
    }
    .stDownloadButton > button {
        width: 100%;
        background-color: #198754;
        color: white;
        font-weight: 600;
        border-radius: 6px;
        padding: 0.4rem 0.8rem;
        border: none;
    }
    </style>
""",
    unsafe_allow_html=True,
)

geod = Geod(ellps="WGS84")


def convert_sheet_name_to_arabic(sheet_str):
  roman_map = {"IV": "4", "III": "3", "II": "2", "I": "1"}
  clean_str = sheet_str.strip()

  for roman, arabic in roman_map.items():
    pattern = rf"[\s\-_]+{roman}$"
    if re.search(pattern, clean_str, re.IGNORECASE):
      clean_str = re.sub(pattern, arabic, clean_str, flags=re.IGNORECASE)
      return clean_str

  return clean_str.replace(" ", "").replace("-", "")


def dd_to_dms(dd, is_latitude=True):
  direction = (
      ("N" if dd >= 0 else "S") if is_latitude else ("E" if dd >= 0 else "W")
  )
  dd = abs(dd)
  degrees = int(dd)
  minutes_decimal = (dd - degrees) * 60
  minutes = int(minutes_decimal)
  seconds = (minutes_decimal - minutes) * 60
  return f"{degrees}° {minutes}' {seconds:.2f}\" {direction}"


def calculate_geodesic_offset(xmin, ymin, xmax, ymax, offset_ns_nm, offset_ew_nm):
  dist_y_m = offset_ns_nm * 1852.0
  dist_x_m = offset_ew_nm * 1852.0

  lon_nw, lat_nw, _ = geod.fwd(xmin, ymax, 270, dist_x_m)
  lon_nw, lat_nw, _ = geod.fwd(lon_nw, lat_nw, 0, dist_y_m)

  lon_ne, lat_ne, _ = geod.fwd(xmax, ymax, 90, dist_x_m)
  lon_ne, lat_ne, _ = geod.fwd(lon_ne, lat_ne, 0, dist_y_m)

  lon_se, lat_se, _ = geod.fwd(xmax, ymin, 90, dist_x_m)
  lon_se, lat_se, _ = geod.fwd(lon_se, lat_se, 180, dist_y_m)

  lon_sw, lat_sw, _ = geod.fwd(xmin, ymin, 270, dist_x_m)
  lon_sw, lat_sw, _ = geod.fwd(lon_sw, lat_sw, 180, dist_y_m)

  return {
      "NW": (lon_nw, lat_nw),
      "NE": (lon_ne, lat_ne),
      "SE": (lon_se, lat_se),
      "SW": (lon_sw, lat_sw),
  }


def generate_excel(csv_rows):
  wb = openpyxl.Workbook()
  ws = wb.active
  ws.title = "NOTAM_Coordinates"

  headers = [
      "Corner Position",
      "Latitude (DMS)",
      "Longitude (DMS)",
      "Latitude (DD)",
      "Longitude (DD)",
  ]
  ws.append(headers)
  for row in csv_rows:
    ws.append(row)

  HEADER_FILLS = {
      "A": PatternFill("solid", fgColor="4682B4"),
      "B": PatternFill("solid", fgColor="4196B4"),
      "C": PatternFill("solid", fgColor="4196B4"),
      "D": PatternFill("solid", fgColor="1F497D"),
      "E": PatternFill("solid", fgColor="1F497D"),
  }
  DATA_FILLS = {
      "A": PatternFill("solid", fgColor="E6F0FA"),
      "B": PatternFill("solid", fgColor="E0F2F1"),
      "C": PatternFill("solid", fgColor="E0F2F1"),
      "D": PatternFill("solid", fgColor="DCE6F1"),
      "E": PatternFill("solid", fgColor="DCE6F1"),
  }

  font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
  font_bold_data = Font(name="Calibri", size=11, bold=True)
  font_regular_data = Font(name="Calibri", size=11)
  align_center = Alignment(horizontal="center", vertical="center")
  white_border = Border(
      left=Side(style="thin", color="FFFFFF"),
      right=Side(style="thin", color="FFFFFF"),
      top=Side(style="thin", color="FFFFFF"),
      bottom=Side(style="thin", color="FFFFFF"),
  )

  for col_idx, cell in enumerate(ws[1], start=1):
    col_letter = openpyxl.utils.get_column_letter(col_idx)
    cell.fill = HEADER_FILLS[col_letter]
    cell.font = font_header
    cell.alignment = align_center
    cell.border = white_border

  for row in ws.iter_rows(min_row=2, max_row=len(csv_rows) + 1, max_col=5):
    for cell in row:
      col_letter = openpyxl.utils.get_column_letter(cell.column)
      cell.fill = DATA_FILLS[col_letter]
      cell.alignment = align_center
      cell.border = white_border
      cell.font = (
          font_bold_data if col_letter == "A" else font_regular_data
      )
      if isinstance(cell.value, float):
        cell.number_format = "0.000000"

  for col in ws.columns:
    max_len = max(len(str(cell.value or "")) for cell in col)
    col_letter = openpyxl.utils.get_column_letter(col[0].column)
    ws.column_dimensions[col_letter].width = max(max_len + 5, 18)

  output = io.BytesIO()
  wb.save(output)
  return output.getvalue()


@st.cache_data
def load_index():
  filename = None
  if os.path.exists("index_l7018.geojson"):
    filename = "index_l7018.geojson"
  elif os.path.exists("index_l7018.json"):
    filename = "index_l7018.json"

  if filename:
    gdf = gpd.read_file(filename)
    if gdf.crs is not None and gdf.crs.to_epsg() != 4326:
      gdf = gdf.to_crs(epsg=4326)
    return gdf
  return None


gdf_index = load_index()

# 📌 SIDEBAR CONFIGURATION
with st.sidebar:
  st.subheader("⚙️ ตั้งค่าบล็อกงาน")

  if gdf_index is not None:
    target_col = "Sheet_dash"
    matched_cols = [
        col for col in gdf_index.columns if col.lower() == target_col.lower()
    ]
    selected_col = (
        matched_cols[0]
        if matched_cols
        else st.selectbox(
            "เลือก Field เลขระวาง:",
            [
                col
                for col in gdf_index.columns
                if col != gdf_index.geometry.name
            ],
        )
    )

    sheet_list = sorted(gdf_index[selected_col].astype(str).unique())
    default_select = [sheet_list[0]] if sheet_list else []

    selected_sheets = st.multiselect(
        "เลือกระวาง L7018:",
        options=sheet_list,
        default=default_select,
        help="พิมพ์ค้นหาหมายเลขระวางได้ที่นี่ครับ",
    )
  else:
    st.error("❌ ไม่พบไฟล์ index_l7018.geojson")
    selected_sheets = []

  arabic_sheets = [
      convert_sheet_name_to_arabic(s) for s in (selected_sheets or [])
  ]

  default_folder_name = (
      f"NOTAM_{'_'.join(arabic_sheets)}" if arabic_sheets else "NOTAM_BLOCK"
  )
  if len(default_folder_name) > 30:
    default_folder_name = f"NOTAM_BLOCK_{len(selected_sheets)}SHEETS"

  st.markdown("---")
  st.subheader("📏 BUFFER")
  folder_name = st.text_input(
      "ชื่อโฟลเดอร์ / ชื่อ NOTAM:", default_folder_name
  )

  col_b1, col_b2 = st.columns(2)
  with col_b1:
    offset_ns = st.number_input(
        "N-S (NM):", min_value=0.0, value=0.5, step=0.1
    )
  with col_b2:
    offset_ew = st.number_input(
        "E-W (NM):", min_value=0.0, value=0.5, step=0.1
    )

  st.markdown("<br>", unsafe_allow_html=True)
  btn_generate = st.button("🚀 ประมวลผลสร้าง Package")

  download_container = st.container()

# 📌 MAIN CONTENT AREA
if gdf_index is not None and len(selected_sheets) > 0:
  selected_polys = gdf_index[
      gdf_index[selected_col].astype(str).isin(selected_sheets)
  ]
  merged_poly = selected_polys.union_all()

  bounds = selected_polys.total_bounds
  xmin, ymin, xmax, ymax = bounds
  new_corners = calculate_geodesic_offset(
      xmin, ymin, xmax, ymax, offset_ns, offset_ew
  )

  kml_sheet_name = (
      f"{'_'.join(arabic_sheets)}.kml" if arabic_sheets else "sheets.kml"
  )

  # 📌 HEADER & TOP DETAILS LAYOUT (รูปภาพ: สี่เหลี่ยมสีแดง + สีส้ม)
  col_hdr_left, col_hdr_right = st.columns([1.4, 1.0])

  with col_hdr_left:
    st.markdown(
        '<h3 class="header-title">✈️ NOTAM AREA GENERATOR</h3>',
        unsafe_allow_html=True,
    )
    st.markdown(
        '<p class="header-desc">ระบบสร้างไฟล์ KML'
        " และตารางพิกัดขอบเขต NOTAM"
        " อัตโนมัติสำหรับงานบินถ่ายภาพทางอากาศ</p>",
        unsafe_allow_html=True,
    )

  with col_hdr_right:
    # 🟧 สี่เหลี่ยมสีส้ม: จัดวาง Metric 4 ตัวเป็น 2x2 ขนาดเล็กกะทัดรัด
    st.markdown(
        f"""
        <div class="metric-grid">
            <div class="metric-card-mini">
                <h5>จำนวนระวาง</h5>
                <p>{len(selected_sheets)} Sheet(s)</p>
            </div>
            <div class="metric-card-mini">
                <h5>BUFFER N-S / E-W</h5>
                <p>{offset_ns} / {offset_ew} NM</p>
            </div>
            <div class="metric-card-mini">
                <h5>ไฟล์ KML ระวาง</h5>
                <p title="{kml_sheet_name}">{kml_sheet_name}</p>
            </div>
            <div class="metric-card-mini">
                <h5>NOTAM PACKAGE</h5>
                <p title="{folder_name}">{folder_name}</p>
            </div>
        </div>
        """,
        unsafe_allow_html=True,
    )

  # 📌 🟦 สี่เหลี่ยมสีฟ้า: MAP & DATA TABS (ขยายเต็มพื้นที่ด้านล่าง)
  tab_map, tab_data = st.tabs(
      ["🗺️ พรีวิวแผนที่ (Interactive Map)", "📍 พิกัดมุม Buffer (Coordinates Table)"]
  )

  with tab_map:
    m = folium.Map(
        location=[(ymin + ymax) / 2, (xmin + xmax) / 2],
        zoom_start=9,
        tiles="OpenStreetMap",
    )

    # 4. ⚡ OPTIMIZED LIGHTWEIGHT INDEX LAYER (เส้นจางมาก + บาง 0.5px + Static)
    folium.GeoJson(
        gdf_index,
        name="Index L7018 Light",
        style_function=lambda feature: {
            "fillColor": "#3186cc",
            "color": "#1e40af",
            "weight": 0.5,  # ลดความหนาเส้น
            "fillOpacity": 0.02,  # จางสุดๆ เพื่อประหยัดการ Render
        },
    ).add_to(m)

    # 2. Merged Selected Boundary (ขอบระวางที่เลือก - เส้นสีเหลือง)
    if merged_poly.geom_type == "Polygon":
      m_coords = list(merged_poly.exterior.coords)
      folium.PolyLine(
          [[c[1], c[0]] for c in m_coords],
          color="#FFD700",
          weight=3.5,
          opacity=0.9,
          tooltip=f"Selected Block ({', '.join(arabic_sheets)})",
      ).add_to(m)
    elif merged_poly.geom_type == "MultiPolygon":
      for poly in merged_poly.geoms:
        m_coords = list(poly.exterior.coords)
        folium.PolyLine(
            [[c[1], c[0]] for c in m_coords],
            color="#FFD700",
            weight=3.5,
            opacity=0.9,
        ).add_to(m)

    # 3. NOTAM Buffer Boundary (ขอบเขต Buffer - เส้นสีแดง)
    buf_coords = [
        [new_corners["NW"][1], new_corners["NW"][0]],
        [new_corners["NE"][1], new_corners["NE"][0]],
        [new_corners["SE"][1], new_corners["SE"][0]],
        [new_corners["SW"][1], new_corners["SW"][0]],
        [new_corners["NW"][1], new_corners["NW"][0]],
    ]
    folium.PolyLine(
        buf_coords,
        color="#DC3545",
        weight=2.5,
        opacity=0.9,
        tooltip="NOTAM Buffer Boundary",
    ).add_to(m)

    # 📌 Render Map แบบ Static (ปรับ height ให้สูงเต็มพื้นที่หน้าจอ ~720px)
    st_folium(m, width="100%", height=720, returned_objects=[])

  with tab_data:
    csv_rows = [
        (
            "NW (บน-ซ้าย)",
            dd_to_dms(new_corners["NW"][1], True),
            dd_to_dms(new_corners["NW"][0], False),
            new_corners["NW"][1],
            new_corners["NW"][0],
        ),
        (
            "NE (บน-ขวา)",
            dd_to_dms(new_corners["NE"][1], True),
            dd_to_dms(new_corners["NE"][0], False),
            new_corners["NE"][1],
            new_corners["NE"][0],
        ),
        (
            "SE (ล่าง-ขวา)",
            dd_to_dms(new_corners["SE"][1], True),
            dd_to_dms(new_corners["SE"][0], False),
            new_corners["SE"][1],
            new_corners["SE"][0],
        ),
        (
            "SW (ล่าง-ซ้าย)",
            dd_to_dms(new_corners["SW"][1], True),
            dd_to_dms(new_corners["SW"][0], False),
            new_corners["SW"][1],
            new_corners["SW"][0],
        ),
    ]

    st.markdown("##### 📌 ตารางสรุปพิกัดมุม 4 ทิศทาง (WGS84)")
    table_data = [
        {
            "ตำแหน่ง (Corner)": r[0],
            "Latitude (DMS)": r[1],
            "Longitude (DMS)": r[2],
            "Latitude (DD)": f"{r[3]:.6f}",
            "Longitude (DD)": f"{r[4]:.6f}",
        }
        for r in csv_rows
    ]
    st.dataframe(table_data, use_container_width=True)

  # 📦 Processing & Package Generator
  if btn_generate:
    safe_name = folder_name.replace(" ", "_")
    sheets_filename = "_".join(arabic_sheets)

    kml_orig = simplekml.Kml()
    if merged_poly.geom_type == "Polygon":
      coords = [(c[0], c[1]) for c in merged_poly.exterior.coords]
      pol = kml_orig.newpolygon(name=sheets_filename, outerboundaryis=coords)
      pol.style.polystyle.color = "00000000"
      pol.style.linestyle.color = simplekml.Color.yellow
      pol.style.linestyle.width = 3
    elif merged_poly.geom_type == "MultiPolygon":
      for idx, poly in enumerate(merged_poly.geoms):
        coords = [(c[0], c[1]) for c in poly.exterior.coords]
        pol = kml_orig.newpolygon(
            name=f"{sheets_filename}_{idx+1}", outerboundaryis=coords
        )
        pol.style.polystyle.color = "00000000"
        pol.style.linestyle.color = simplekml.Color.yellow
        pol.style.linestyle.width = 3

    kml_buf = simplekml.Kml()
    pol_buf = kml_buf.newpolygon(
        name=f"{safe_name}_Buffer",
        outerboundaryis=[
            new_corners["NW"],
            new_corners["NE"],
            new_corners["SE"],
            new_corners["SW"],
            new_corners["NW"],
        ],
    )
    pol_buf.style.polystyle.color = "00000000"
    pol_buf.style.linestyle.color = simplekml.Color.red
    pol_buf.style.linestyle.width = 3

    for r in csv_rows:
      pnt = kml_buf.newpoint(name=r[0], coords=[(r[4], r[3])])
      pnt.description = f"Lat: {r[1]}\nLon: {r[2]}"

    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zip_file:
      zip_file.writestr(
          f"{safe_name}/{sheets_filename}.kml", kml_orig.kml()
      )
      zip_file.writestr(f"{safe_name}/{safe_name}.kml", kml_buf.kml())
      zip_file.writestr(
          f"{safe_name}/{safe_name}_coordinates.xlsx",
          generate_excel(csv_rows),
      )

    with download_container:
      st.markdown("<br>", unsafe_allow_html=True)
      st.download_button(
          label="📥 ดาวน์โหลด NOTAM Package (.zip)",
          data=zip_buffer.getvalue(),
          file_name=f"{safe_name}.zip",
          mime="application/zip",
      )

else:
  st.info("👈 กรุณาเลือกหมายเลขระวาง L7018 ที่เมนูด้านซ้ายเพื่อเริ่มต้นใช้งาน")
